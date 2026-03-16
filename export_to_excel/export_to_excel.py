"""
title: Export to Excel
description: Export Assistant Message Tables to Excel
version: 1.0.0
author: Helmi Chaouachi
git_url: https://github.com/Helmi97/open-webui-extensions/tree/main/export_to_excel
required_open_webui_version: 0.8.0
requirements: openpyxl,markdown,beautifulsoup4
"""

from __future__ import annotations

import base64
import io
import json
import re
from typing import Any

import markdown
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from pydantic import BaseModel, Field

INVALID_SHEET_TITLE_RE = re.compile(r"[:\\/?*\[\]]")
FORMULA_PREFIXES = ("=", "+", "-", "@")


class Action:
    class Valves(BaseModel):
        priority: int = Field(
            default=0,
            description="Controls button display order (lower = appears first).",
        )
        filename_prefix: str = Field(
            default="message",
            description="Prefix used for the downloaded Excel file name.",
        )

    def __init__(self):
        self.valves = self.Valves()

    def build_filename(self, message_id: str) -> str:
        return f"{self.valves.filename_prefix}-{message_id}.xlsx"

    async def emit_error(self, message: str, __event_emitter__=None):
        if __event_emitter__:
            await __event_emitter__(
                {
                    "type": "notification",
                    "data": {"type": "error", "content": message},
                }
            )

    async def emit_status(
        self, description: str, done: bool, __event_emitter__=None
    ):
        if __event_emitter__:
            await __event_emitter__(
                {
                    "type": "status",
                    "data": {"description": description, "done": done},
                }
            )

    def _normalize_content(self, content: Any) -> str:
        if content is None:
            return ""

        if isinstance(content, str):
            return content

        if isinstance(content, list):
            parts: list[str] = []
            for item in content:
                if not isinstance(item, dict):
                    continue
                if item.get("type") == "text":
                    parts.append(item.get("text", ""))
                elif isinstance(item.get("content"), str):
                    parts.append(item.get("content"))
            return "\n".join(part for part in parts if part)

        return str(content)

    def get_message_content(self, body: dict) -> str:
        target_id = body.get("id")
        messages = body.get("messages", []) or []

        if target_id:
            for message in messages:
                if not isinstance(message, dict):
                    continue

                candidate_ids = [
                    message.get("id"),
                    message.get("message_id"),
                    (
                        (message.get("info") or {}).get("id")
                        if isinstance(message.get("info"), dict)
                        else None
                    ),
                    (
                        (message.get("meta") or {}).get("id")
                        if isinstance(message.get("meta"), dict)
                        else None
                    ),
                ]

                if target_id in candidate_ids and message.get("role") == "assistant":
                    return self._normalize_content(message.get("content"))

        for message in reversed(messages):
            if isinstance(message, dict) and message.get("role") == "assistant":
                return self._normalize_content(message.get("content"))

        return ""

    def markdown_to_html(self, markdown_text: str) -> str:
        return markdown.markdown(
            markdown_text,
            extensions=["tables", "fenced_code", "sane_lists"],
        )

    def _get_table_rows(self, table: Tag) -> list[tuple[Tag, bool]]:
        rows: list[tuple[Tag, bool]] = []

        for child in table.children:
            if not isinstance(child, Tag):
                continue

            if child.name == "tr":
                rows.append((child, False))
            elif child.name in ("thead", "tbody", "tfoot"):
                is_header_section = child.name == "thead"
                for row in child.find_all("tr", recursive=False):
                    rows.append((row, is_header_section))

        if rows:
            return rows

        return [(row, False) for row in table.find_all("tr")]

    def _sanitize_cell_value(self, value: str) -> str:
        normalized = " ".join(value.split())
        if normalized.lstrip().startswith(FORMULA_PREFIXES):
            return f"'{normalized}"
        return normalized

    def extract_tables(self, markdown_text: str) -> list[dict[str, Any]]:
        soup = BeautifulSoup(self.markdown_to_html(markdown_text), "html.parser")
        tables: list[dict[str, Any]] = []

        for table in soup.find_all("table"):
            parsed_rows: list[list[str]] = []
            header_rows: list[bool] = []

            for row, in_header_section in self._get_table_rows(table):
                cells = row.find_all(["th", "td"], recursive=False)
                if not cells:
                    continue

                values = [
                    self._sanitize_cell_value(cell.get_text(" ", strip=True))
                    for cell in cells
                ]
                parsed_rows.append(values)
                header_rows.append(
                    in_header_section or all(cell.name == "th" for cell in cells)
                )

            if parsed_rows:
                tables.append({"rows": parsed_rows, "header_rows": header_rows})

        return tables

    def _make_sheet_title(self, index: int, existing_titles: set[str]) -> str:
        base_title = f"Table {index}"
        base_title = INVALID_SHEET_TITLE_RE.sub(" ", base_title).strip() or "Table"
        base_title = base_title[:31]

        if base_title not in existing_titles:
            existing_titles.add(base_title)
            return base_title

        suffix = 2
        while True:
            suffix_text = f" ({suffix})"
            candidate = f"{base_title[: 31 - len(suffix_text)]}{suffix_text}".strip()
            if candidate not in existing_titles:
                existing_titles.add(candidate)
                return candidate
            suffix += 1

    def build_workbook(self, tables: list[dict[str, Any]]) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)

        header_font = Font(bold=True)
        wrap_alignment = Alignment(vertical="top", wrap_text=True)
        existing_titles: set[str] = set()

        for index, table in enumerate(tables, start=1):
            worksheet = workbook.create_sheet(
                title=self._make_sheet_title(index, existing_titles)
            )
            rows = table["rows"]
            header_rows = table["header_rows"]

            max_lengths: dict[int, int] = {}
            header_prefix_count = 0
            for is_header_row in header_rows:
                if not is_header_row:
                    break
                header_prefix_count += 1

            for row_index, row in enumerate(rows, start=1):
                worksheet.append(row)
                for column_index, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_index, column=column_index)
                    cell.alignment = wrap_alignment
                    max_lengths[column_index] = max(
                        max_lengths.get(column_index, 0), len(value)
                    )
                    if header_rows[row_index - 1]:
                        cell.font = header_font

            if header_prefix_count > 0 and header_prefix_count < len(rows):
                worksheet.freeze_panes = f"A{header_prefix_count + 1}"

            for column_index, max_length in max_lengths.items():
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=column_index).column_letter
                ].width = min(max(max_length + 2, 10), 60)

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    async def download_file(
        self,
        xlsx_bytes: bytes,
        filename: str,
        __event_emitter__=None,
        __event_call__=None,
    ):
        encoded = base64.b64encode(xlsx_bytes).decode("ascii")

        js_code = f"""
const base64 = {json.dumps(encoded)};
const filename = {json.dumps(filename)};
const binary = atob(base64);
const bytes = new Uint8Array(binary.length);

for (let i = 0; i < binary.length; i++) {{
  bytes[i] = binary.charCodeAt(i);
}}

const blob = new Blob(
  [bytes],
  {{ type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" }}
);
const url = URL.createObjectURL(blob);

try {{
  const a = document.createElement("a");
  a.href = url;
  a.download = filename;
  a.style.display = "none";
  document.body.appendChild(a);
  a.click();
  a.remove();
}} finally {{
  setTimeout(() => URL.revokeObjectURL(url), 4000);
}}

return {{ success: true, filename, size: bytes.length }};
"""

        payload = {"type": "execute", "data": {"code": js_code}}

        if __event_call__ is not None:
            return await __event_call__(payload)

        if __event_emitter__ is not None:
            await __event_emitter__(payload)
            return {"success": True, "filename": filename, "size": len(xlsx_bytes)}

        return None

    async def action(
        self,
        body: dict,
        __event_emitter__=None,
        __event_call__=None,
        **kwargs,
    ):
        message_id = body.get("id")
        if not message_id:
            await self.emit_error(
                "Excel export failed: could not determine the current message id.",
                __event_emitter__,
            )
            return {
                "content": "Could not determine the current message id from body['id']."
            }

        filename = self.build_filename(message_id)

        await self.emit_status("Reading message content...", False, __event_emitter__)
        markdown_text = self.get_message_content(body)
        if not markdown_text.strip():
            await self.emit_error(
                "Excel export failed: no assistant message content found.",
                __event_emitter__,
            )
            await self.emit_status(
                "No assistant message content found.",
                True,
                __event_emitter__,
            )
            return {"content": "No assistant message content found."}

        await self.emit_status("Extracting tables...", False, __event_emitter__)
        tables = self.extract_tables(markdown_text)
        if not tables:
            await self.emit_error(
                "Excel export failed: no tables found in the assistant message.",
                __event_emitter__,
            )
            await self.emit_status(
                "No tables found in the assistant message.",
                True,
                __event_emitter__,
            )
            return {"content": "No tables found in the assistant message."}

        await self.emit_status("Generating Excel file...", False, __event_emitter__)
        try:
            xlsx_bytes = self.build_workbook(tables)
        except Exception as exc:
            await self.emit_error(f"Excel export failed: {exc}", __event_emitter__)
            await self.emit_status("Excel export failed.", True, __event_emitter__)
            return {"content": f"Excel export failed: {exc}"}

        await self.emit_status("Starting download...", False, __event_emitter__)
        result = await self.download_file(
            xlsx_bytes=xlsx_bytes,
            filename=filename,
            __event_emitter__=__event_emitter__,
            __event_call__=__event_call__,
        )

        await self.emit_status("Excel export complete.", True, __event_emitter__)
        return {
            "content": f"Exported message tables to Excel: {filename}",
            "result": result,
            "table_count": len(tables),
        }
